import os
import re
import csv
from io import StringIO
from datetime import datetime
import httpx
import shutil

import pdfplumber
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()


# =========================================================
# CONFIG / PROMPT
# =========================================================

EXPECTED_HEADERS = [
    "Archivo",
    "BaseImp",
    "BaseIRPF",
    "Buyer",
    "Empresa",
    "FEscaneo",
    "FFactura",
    "FOperacion",
    "ImporIVA",
    "Moneda",
    "NombreProveedor",
    "NumeroFactura",
    "PedidoCliente",
    "Proveedor",
    "TipoIVA",
    "TipoIVA2",
    "TipoIVA3",
    "TotalFact",
]

DEFAULT_PROMPT = """Eres un extractor estricto de datos de facturas de proveedor.

Debes devolver SOLO una tabla de texto separada por |.

NO uses comas como separador.
NO uses markdown.
NO des explicaciones.
NO añadas texto fuera de la tabla.
NO uses saltos de línea dentro de ninguna celda.
Cada celda debe contener una sola línea de texto.

La salida debe tener exactamente 2 líneas:
- 1 línea de cabecera
- 1 línea de datos

La cabecera debe ser EXACTAMENTE:

Archivo|BaseImp|BaseIRPF|Buyer|Empresa|FEscaneo|FFactura|FOperacion|ImporIVA|Moneda|NombreProveedor|NumeroFactura|PedidoCliente|Proveedor|TipoIVA|TipoIVA2|TipoIVA3|TotalFact

Reglas generales:
- Si un dato no aparece claramente, devuelve "-".
- No inventes ningún dato.
- No deduzcas datos que no estén explícitos en la factura.
- Mantén el texto limpio, sin saltos de línea.
- No añadas comentarios.
- No añadas unidades ni símbolos de moneda salvo que formen parte inseparable del dato.
- Para importes, devuelve solo el número, usando coma decimal si aparece así en el documento.
- Para fechas, devuelve el formato que aparezca en el documento. Si puedes normalizar con seguridad, usa DD/MM/AAAA.

Definición de campos:

IDFactura:
- Identificador único de la factura.
- Normalmente coincide con el ID documental o identificador principal de la factura.
- Si no aparece claramente devuelve "-".

BaseImp:
- Base imponible total de la factura.
- Devuelve únicamente el importe.

BaseIRPF:
- Base del IRPF si existe.
- Si no existe devuelve "-".

Buyer:
- CIF/NIF/VAT del comprador.
- Devuelve únicamente el identificador fiscal.
- Nunca devuelvas el nombre de la empresa.

Empresa:
- Nombre de la empresa compradora.
- Si existe un CIF/NIF/VAT del comprador (Buyer), busca también la razón social asociada.
- No devuelvas "-" si la empresa compradora aparece identificada en cualquier parte de la factura.

FEscaneo:
- Siempre devuelve "-".
- Este campo será completado posteriormente por el sistema.

FFactura:
- Fecha de factura.

FOperacion:
- Si no aparece claramente devuelve "-".

ImporIVA:
- Si la factura indica que no existe IVA, VAT o impuesto aplicable, devuelve 0.

Moneda:
- Devuelve EUR, USD, GBP o la moneda indicada en la factura.
- Si aparece € devuelve EUR.

NombreProveedor:
- Razón social del proveedor.

NumeroFactura:
- Número de factura del proveedor.

PedidoCliente:
- Número de pedido del cliente.
- Puede aparecer como:
  Pedido Cliente
  Customer Order
  Customer PO
  Purchase Order
  PO Number
  Order Number
  Nº Pedido
  Pedido
  Order Ref
  Customer Reference
- NO devolver el número de factura.
- Devuelve únicamente el pedido del cliente.
- Si no existe devuelve "-".

Proveedor:
- CIF/NIF/VAT del proveedor.
- Devuelve únicamente el identificador fiscal.

TipoIVA:
- Si la factura indica exención, inversión del sujeto pasivo o impuesto 0%, devuelve 0.
- No devuelvas "-" cuando pueda determinarse que el IVA es cero.

TipoIVA2:
- Segundo tipo IVA si existe.

TipoIVA3:
- Tercer tipo IVA si existe.

TotalFact:
- Importe total final de la factura.


IMPORTANTE:

Las facturas pueden estar en cualquier idioma.

Debes reconocer automáticamente los campos aunque aparezcan en:
- español
- inglés
- portugués
- francés
- italiano
- alemán
- neerlandés
- chino
- griego
- otros idiomas

No dependas del idioma para localizar la información.

Identifica los conceptos por su significado semántico y no por palabras exactas.
"""


# =========================================================
# CLIENTE OPENAI / AZURE
# =========================================================

import httpx
from openai import OpenAI

def paginas_pdf(pdf_path):
    
    paginas = []

    with pdfplumber.open(pdf_path) as pdf:

        for page in pdf.pages:

            texto = page.extract_text() or ""

            paginas.append(texto)

    return paginas


def detectar_numero_factura(texto):
    
    client = build_client()

    response = client.chat.completions.create(
        model=get_model(),
        temperature=0,
        messages=[
            {
                "role": "system",
                "content": """
Devuelve únicamente el número de factura.

Si no existe devuelve -.

Sin explicaciones.
"""
            },
            {
                "role": "user",
                "content": texto[:15000]
            }
        ]
    )

    return response.choices[0].message.content.strip()

def agrupar_facturas(pdf_path):
    
    paginas = paginas_pdf(pdf_path)

    grupos = {}

    for pagina in paginas:

        numero = detectar_numero_factura(pagina)

        if numero not in grupos:
            grupos[numero] = []

        grupos[numero].append(pagina)

    return [
        "\n".join(paginas_factura)
        for paginas_factura in grupos.values()
    ]


def build_client():
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    base_url = os.getenv("OPENAI_BASE_URL", "").strip()

    if not api_key:
        raise Exception("No se ha encontrado OPENAI_API_KEY. Revisa el .env")

    http_client = httpx.Client(verify=False)

    if base_url:
        return OpenAI(
            api_key=api_key,
            base_url=base_url,
            http_client=http_client
        )

    return OpenAI(
        api_key=api_key,
        http_client=http_client
    )


def get_model():
    return os.getenv("OPENAI_MODEL", "gpt-4.1").strip()


# =========================================================
# PDF
# =========================================================

def read_pdf_text(path):
    
    text = ""

    with pdfplumber.open(path) as pdf:

        paginas = pdf.pages

        if len(paginas) > 10:

            seleccion = (
                paginas[:2]
                + paginas[-5:]
            )

        else:
            seleccion = paginas

        for page in seleccion:

            content = page.extract_text()

            if content:
                text += content + "\n"

    return text.strip()


def clean_pdf_text(text):
    if not text:
        return ""

    text = text.replace("\x00", " ")
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()


def pdf_a_markdown_factura(text):
    """
    Convierte el texto extraído de la factura en pseudo-markdown
    para ayudar al modelo a localizar bloques relevantes.
    """
    if not text:
        return ""

    text = clean_pdf_text(text)
    lineas = [l.strip() for l in text.splitlines() if l.strip()]
    resultado = []

    claves_seccion = [
        "FACTURA",
        "INVOICE",
        "PROVEEDOR",
        "SUPPLIER",
        "VENDOR",
        "CLIENTE",
        "CUSTOMER",
        "NIF",
        "CIF",
        "VAT",
        "BASE IMPONIBLE",
        "BASE",
        "IVA",
        "VAT",
        "TOTAL",
        "FECHA",
        "DATE",
        "VENCIMIENTO",
        "DUE DATE",
        "DIRECCIÓN",
        "DIRECCION",
        "ADDRESS",
        "POSTAL",
        "COUNTRY",
        "PAÍS",
        "PAIS",
    ]

    for linea in lineas:
        up = linea.upper()

        if any(k in up for k in claves_seccion):
            resultado.append(f"\n## {linea}\n")
        else:
            resultado.append(linea)

    md = "\n".join(resultado)
    md = re.sub(r"\n{3,}", "\n\n", md).strip()

    return md


# =========================================================
# CSV / UTILIDADES
# =========================================================

def csv_to_matrix(csv_text):
    lines = [line.strip() for line in str(csv_text).splitlines() if line.strip()]
    return [line.split("|") for line in lines]


def clean_llm_csv_response(text):
    if not text:
        return ""

    text = text.strip()
    text = re.sub(r"^```(?:csv|text)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text)

    expected_header = "|".join(EXPECTED_HEADERS)
    idx = text.find(expected_header)

    if idx != -1:
        text = text[idx:].strip()

    lines = [line.strip() for line in text.splitlines() if line.strip()]

    if len(lines) >= 2:
        lines = lines[:2]

    cleaned_lines = []

    for line in lines:
        cols = line.split("|")
        cols = (cols + [""] * len(EXPECTED_HEADERS))[:len(EXPECTED_HEADERS)]
        cleaned_lines.append("|".join(cols))

    return "\n".join(cleaned_lines)


def normalizar_valor(v):
    v = str(v).strip()

    if v in ["", "N/A", "NA", "No aplica", "NO APLICA", "n/a", "na", "None", "null"]:
        return "-"

    return v


def limpiar_fila(row):
    row = (row + [""] * len(EXPECTED_HEADERS))[:len(EXPECTED_HEADERS)]
    return [normalizar_valor(x) for x in row]


def combinar_csvs(lista_csv):
    """
    Une varios CSV individuales en una única tabla.
    Mantiene una sola cabecera.
    """
    todas = []

    for csv_text in lista_csv:
        rows = csv_to_matrix(csv_text)

        if len(rows) < 2:
            continue

        if not todas:
            todas.append(EXPECTED_HEADERS)

        todas.append(limpiar_fila(rows[1]))

    output = StringIO()
    writer = csv.writer(output, delimiter="|", lineterminator="\n")

    for row in todas:
        writer.writerow(row)

    return output.getvalue().strip()


# =========================================================
# LLAMADA AL MODELO
# =========================================================

def extract_invoice_with_agent(file_name, invoice_text, agent_prompt=DEFAULT_PROMPT):
    client = build_client()
    model = get_model()

    raw_text = clean_pdf_text(invoice_text)
    llm_text = pdf_a_markdown_factura(raw_text)
    MAX_CHARS = 35000
    if len(raw_text) > MAX_CHARS:
        
        mitad = MAX_CHARS // 2

        raw_text = (
                raw_text[:mitad]
                + "\n\n"
                + raw_text[-mitad:]
            )

        
    final_prompt = f"""
{agent_prompt}

NOMBRE DEL ARCHIVO:
\"\"\"
{file_name}
\"\"\"

FACTURA:
\"\"\"
{llm_text}
\"\"\"
"""

    response = client.chat.completions.create(
        model=model,
        temperature=0,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        messages=[
            {
                "role": "system",
                "content": (
                    "Eres un extractor documental muy estricto. "
                    "Tu única salida permitida es una tabla válida separada por |, "
                    "sin explicaciones, sin markdown y sin texto adicional."
                )
            },
            {
                "role": "user",
                "content": final_prompt
            }
        ]
    )

    raw = response.choices[0].message.content or ""
    cleaned = clean_llm_csv_response(raw)

    rows = csv_to_matrix(cleaned)

    if len(rows) < 2:
        fixed = "|".join(EXPECTED_HEADERS) + "\n" + "|".join([file_name] + ["-"] * (len(EXPECTED_HEADERS) - 1))
        return fixed

    headers = rows[0]
    data = limpiar_fila(rows[1])

    if headers != EXPECTED_HEADERS:
        headers = EXPECTED_HEADERS

    # Archivo
    data[0] = file_name

    # FEscaneo
    data[5] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # FOperacion = FFactura
    if data[6] != "-":
        data[7] = data[6]

    # IVA a 0
    if data[9] == "-" and data[2] != "-" and data[17] != "-":
        try:
            base = float(data[2].replace(".", "").replace(",", "."))
            total = float(data[17].replace(".", "").replace(",", "."))

            if abs(base - total) < 0.01:
                data[9] = "0"
        except:
            pass

    # Tipo IVA a 0
    if data[14] == "-" and data[9] == "0":
        data[14] = "0"

    # Moneda
    if data[10] == "-":
        texto_up = raw_text.upper()

        if "€" in raw_text or "EUR" in texto_up:
            data[10] = "EUR"

        elif "$" in raw_text or "USD" in texto_up:
            data[10] = "USD"

    # Empresa
    if data[5] == "-" and data[4] != "-":
        # evitar que Buyer se use como Empresa
        pass

    # Buyer
    if data[4] != "-":
        data[4] = data[4].replace(" ", "").upper()

    # Proveedor
    if data[13] != "-":
        data[13] = data[13].replace(" ", "").upper()

    output = StringIO()
    writer = csv.writer(output, delimiter="|", lineterminator="\n")
    writer.writerow(headers)
    writer.writerow(data)

    return output.getvalue().strip()


# =========================================================
# EXCEL EXPORT
# =========================================================

def export_to_excel(csv_text, path):
    rows = csv_to_matrix(csv_text)

    if len(rows) < 2:
        raise ValueError("No hay datos válidos para exportar.")

    headers = rows[0]
    data_rows = rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    fill_header = PatternFill("solid", fgColor="DCE6F1")
    fill_empty = PatternFill("solid", fgColor="E7E6E6")

    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True, color="0F2D52")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(data_rows, start=2):
        expanded = (row + [""] * len(headers))[:len(headers)]

        for col_idx, value in enumerate(expanded, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            if value == "-":
                cell.fill = fill_empty
                cell.font = Font(color="666666")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    last_col = ws.cell(row=1, column=len(headers)).column_letter
    last_row = len(data_rows) + 1

    table = Table(
        displayName="TablaFacturas",
        ref=f"A1:{last_col}{last_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(path)


# =========================================================
# HISTORIAL
# =========================================================

def guardar_historial(csv_text, user_id):
    carpeta = "historial"
    os.makedirs(carpeta, exist_ok=True)

    rows = csv_to_matrix(csv_text)

    if not rows or len(rows) < 2:
        print("⚠ CSV inválido, no se guarda historial")
        return

    fecha = datetime.now().strftime("%Y-%m-%d")
    hora = datetime.now().strftime("%H:%M:%S")
    path = f"{carpeta}/historial_facturas_{user_id}_{fecha}.xlsx"

    headers = rows[0]
    data_rows = rows[1:]

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"

    fill_bloque = PatternFill("solid", fgColor="BDD7EE")
    fill_header = PatternFill("solid", fgColor="DCE6F1")
    fill_empty = PatternFill("solid", fgColor="E7E6E6")

    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    start_row = ws.max_row + 1 if ws.max_row > 1 or ws["A1"].value else 1

    titulo = f"Extracción facturas | Fecha: {fecha} | Hora: {hora} | Nº facturas: {len(data_rows)}"
    ws.cell(row=start_row, column=1, value=titulo)

    total_cols = len(headers)

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=total_cols
    )

    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True, color="0F2D52")
    title_cell.fill = fill_bloque
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = start_row + 1

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = Font(bold=True, color="0F2D52")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    current_row = header_row + 1

    for row in data_rows:
        expanded = (row + [""] * len(headers))[:len(headers)]

        for col_idx, value in enumerate(expanded, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            if value == "-":
                cell.fill = fill_empty
                cell.font = Font(color="666666")

        current_row += 1

    current_row += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(path)

def guardar_factura_corregida(
    archivo,
    columna,
    valor,
    usuario
):
    from datetime import datetime

    path = "historial/facturas_corregidas.xlsx"

    if not os.path.exists(path):

        wb = Workbook()
        ws = wb.active

        ws.append([
            "Archivo",
            *EXPECTED_HEADERS[1:],
            "UsuarioUltimaModificacion",
            "FechaUltimaModificacion"
        ])

        wb.save(path)

    wb = load_workbook(path)
    ws = wb.active

    headers = [
        str(c.value).strip()
        for c in ws[1]
    ]

    fila_pdf = None

    for fila in range(2, ws.max_row + 1):

        if str(ws.cell(fila, 1).value or "").strip() == archivo:

            fila_pdf = fila

    if fila_pdf is None:

        fila_pdf = ws.max_row + 1

        ws.cell(fila_pdf, 1).value = archivo

    if columna in headers:

        col_idx = headers.index(columna) + 1

        ws.cell(
            fila_pdf,
            col_idx
        ).value = valor

    ws.cell(
        fila_pdf,
        len(headers) - 1
    ).value = usuario

    ws.cell(
        fila_pdf,
        len(headers)
    ).value = datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )

    wb.save(path)

def guardar_factura_corregida_completa(
    fila_completa,
    usuario
):


    print("FILA COMPLETA RECIBIDA:")
    print(fila_completa)
    
    from datetime import datetime
    import os

    path = "historial/facturas_corregidas.xlsx"

    headers = (
        EXPECTED_HEADERS
        + [
            "UsuarioUltimaModificacion",
            "FechaUltimaModificacion"
        ]
    )

    if not os.path.exists(path):

        wb = Workbook()
        ws = wb.active
        ws.title = "FacturasCorregidas"

        ws.append(headers)

        wb.save(path)

    wb = load_workbook(path)
    ws = wb.active

    archivo = str(fila_completa[0]).strip()

    fila_objetivo = None

    for fila in range(2, ws.max_row + 1):

        archivo_excel = str(
            ws.cell(fila, 1).value or ""
        ).strip()

        if archivo_excel == archivo:

            fila_objetivo = fila
            break

    if fila_objetivo is None:

        fila_objetivo = ws.max_row + 1

    fila_completa = [
        "-" if str(x).strip() == "" else str(x).strip()
        for x in fila_completa
    ]

    while len(fila_completa) < len(EXPECTED_HEADERS):
        fila_completa.append("-")

    fila_completa = fila_completa[:len(EXPECTED_HEADERS)]

    for col, valor in enumerate(fila_completa, start=1):

        ws.cell(
            row=fila_objetivo,
            column=col
        ).value = valor

    ws.cell(
        row=fila_objetivo,
        column=len(EXPECTED_HEADERS) + 1
    ).value = usuario

    ws.cell(
        row=fila_objetivo,
        column=len(EXPECTED_HEADERS) + 2
    ).value = datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )

    wb.save(path)


def clasificar_factura(fila):
    
    """
    Devuelve:

    procesada
    imagen
    manual
    examinada
    """

    archivo = str(fila[0]).strip()

    # ==========================================
    # YA EXAMINADA
    # ==========================================

    path_corregidas = "historial/facturas_corregidas.xlsx"

    if os.path.exists(path_corregidas):

        wb = load_workbook(path_corregidas)
        ws = wb.active

        for row in range(2, ws.max_row + 1):

            archivo_excel = str(
                ws.cell(row, 1).value or ""
            ).strip()

            if archivo_excel == archivo:

                return "examinada"

    # ==========================================
    # CAMPOS OBLIGATORIOS
    # ==========================================

    obligatorios = [
        "BaseImp",
        "Buyer",
        "Empresa",
        "FFactura",
        "ImporIVA",
        "Moneda",
        "NombreProveedor",
        "NumeroFactura",
        "PedidoCliente",
        "Proveedor",
        "TipoIVA",
        "TotalFact",
    ]

    datos = dict(
        zip(
            EXPECTED_HEADERS,
            fila
        )
    )

    # ==========================================
    # IMAGEN
    # ==========================================

    todos_vacios = True

    for campo in obligatorios:

        valor = str(
            datos.get(campo, "-")
        ).strip()

        if valor != "-":

            todos_vacios = False
            break

    if todos_vacios:

        return "imagen"

    # ==========================================
    # MANUAL
    # ==========================================

    for campo in obligatorios:

        valor = str(
            datos.get(campo, "-")
        ).strip()

        if valor == "-":

            return "manual"

    # ==========================================
    # OK
    # ==========================================

    return "procesada"


def mover_pdf(pdf_path, tipo):
    
    carpetas = {
        "procesada": "facturas/procesadas",
        "manual": "facturas/corregir_manualmente",
        "imagen": "facturas/imagenes",
        "examinada": "facturas/examinadas"
    }

    carpeta_destino = carpetas.get(
        tipo,
        "facturas/error"
    )

    os.makedirs(
        carpeta_destino,
        exist_ok=True
    )

    destino = os.path.join(
        carpeta_destino,
        os.path.basename(pdf_path)
    )

    shutil.copy2(
        pdf_path,
        destino
    )

    print("COPIADO A:", destino)