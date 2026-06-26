import os
import tempfile
from datetime import datetime
from typing import List

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel
from app.logic import guardar_factura_corregida

from app.logic import (
    read_pdf_text,
    extract_invoice_with_agent,
    combinar_csvs,
    csv_to_matrix,
    export_to_excel,
    guardar_historial,
    agrupar_facturas,
    guardar_factura_corregida_completa,
    clasificar_factura,
    mover_pdf, 
)
try:
    from app.Usuario import get_user_id, get_historial_path, leer_historial
except Exception:
    get_user_id = None
    get_historial_path = None
    leer_historial = None


app = FastAPI(title="Extractor de facturas")
class Correccion(BaseModel):
    archivo: str
    columna: str
    valor_anterior: str
    valor_nuevo: str
    usuario: str
    fila_completa: list
# =========================================================
# FRONTEND ESTÁTICO
# =========================================================

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


# =========================================================
# CORS
# =========================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES_DIR = os.path.join(ROOT_DIR, "templates")
STATIC_DIR = os.path.join(ROOT_DIR, "static")

@app.get("/")
def home():
    return FileResponse(os.path.join(TEMPLATES_DIR, "index.html"))


# =========================================================
# EXTRAER FACTURAS
# =========================================================

@app.post("/extraer")
async def extraer_facturas(
    request: Request,
    facturas: List[UploadFile] = File(...)
):
    temp_paths = []

    try:
        if not facturas:
            raise HTTPException(status_code=400, detail="No se han recibido facturas.")

        lista_csv = []
        pdfs_a_mover = []


        for factura in facturas:
            if not factura.filename.lower().endswith(".pdf"):
                raise HTTPException(status_code=400, detail=f"El archivo {factura.filename} no es PDF.")

            os.makedirs("facturas/entrada", exist_ok=True)

            pdf_original = os.path.join(
                "facturas/entrada",
                factura.filename
            )

            contenido = await factura.read()

            with open(pdf_original, "wb") as f:
                f.write(contenido)
                f.flush()
                os.fsync(f.fileno())
                
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(contenido)
                pdf_path = tmp.name
                temp_paths.append(pdf_path)

            text = read_pdf_text(pdf_path)

            # Limitar PDFs enormes
            MAX_CHARS = 35000

            if len(text) > MAX_CHARS:
                text = text[:MAX_CHARS]

            result_csv = extract_invoice_with_agent(
                file_name=factura.filename,
                invoice_text=text
            )

            tabla_tmp = csv_to_matrix(result_csv)

            if len(tabla_tmp) > 1:

                fila = tabla_tmp[1]

                tipo = clasificar_factura(fila)

                print(
                    factura.filename,
                    "=>",
                    tipo
                )

                pdfs_a_mover.append(
                    (pdf_original, tipo)
                )

            lista_csv.append(result_csv)

        result_csv_final = combinar_csvs(lista_csv)
        result_table = csv_to_matrix(result_csv_final)

    

        response = JSONResponse({
            "csv": result_csv_final,
            "tabla": result_table
        })

        guardar_historial(
            result_csv_final,
            user_id="usuario"
        )

        for pdf_original, tipo in pdfs_a_mover:
    
            try:

                mover_pdf(
                    pdf_original,
                    tipo
                )

            except Exception as e:

                print(
                    "ERROR MOVIENDO:",
                    pdf_original,
                    str(e)
                )
        return response

    except Exception as e:
        import traceback
        print("\n❌ ERROR BACKEND:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        for path in temp_paths:
            if path and os.path.exists(path):
                os.remove(path)


# =========================================================
# DESCARGAR EXCEL INDIVIDUAL
# =========================================================

@app.post("/extraer-excel")
async def extraer_y_descargar_excel(
    facturas: List[UploadFile] = File(...)
):
    temp_paths = []

    try:
        if not facturas:
            raise HTTPException(status_code=400, detail="No se han recibido facturas.")

        lista_csv = []
        for factura in facturas:
            if not factura.filename.lower().endswith(".pdf"):
                raise HTTPException(status_code=400, detail=f"El archivo {factura.filename} no es PDF.")

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(await factura.read())
                pdf_path = tmp.name
                temp_paths.append(pdf_path)

            text = read_pdf_text(pdf_path)

            result_csv = extract_invoice_with_agent(
                file_name=factura.filename,
                invoice_text=text
            )

            lista_csv.append(result_csv)

        result_csv_final = combinar_csvs(lista_csv)

        excel_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        export_to_excel(result_csv_final, excel_path)

        return FileResponse(
            excel_path,
            filename="extraccion_facturas.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        print("\n❌ ERROR BACKEND EXCEL:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        for path in temp_paths:
            if path and os.path.exists(path):
                os.remove(path)


# =========================================================
# HISTORIAL JSON PRIVADO
# =========================================================

@app.get("/historial-json")
def obtener_historial():

    fecha = datetime.now().strftime("%Y-%m-%d")

    path = f"historial/historial_facturas_usuario_{fecha}.xlsx"

    if not os.path.exists(path):
        return {"tabla": []}

    wb = load_workbook(path)
    ws = wb.active

    tabla = []

    for row in ws.iter_rows(values_only=True):
        fila = [str(c) if c is not None else "" for c in row]
        tabla.append(fila)

    return {"tabla": tabla}


# =========================================================
# DESCARGAR HISTORIAL PRIVADO
# =========================================================

@app.get("/historial")
def descargar_historial():

    fecha = datetime.now().strftime("%Y-%m-%d")

    path = f"historial/historial_facturas_usuario_{fecha}.xlsx"

    if not os.path.exists(path):
        raise HTTPException(
            status_code=404,
            detail="No existe historial"
        )

    return FileResponse(
        path,
        filename=f"historial_facturas_{fecha}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.post("/guardar-correccion")
async def guardar_correccion(c: Correccion):

    fecha = datetime.now().strftime("%Y-%m-%d")

    path = f"historial/historial_facturas_usuario_{fecha}.xlsx"

    print("================================")
    print("CORRECCION RECIBIDA")
    print("Archivo recibido:", c.archivo)
    print("Columna recibida:", c.columna)
    print("Valor nuevo:", c.valor_nuevo)
    print("================================")

    if not os.path.exists(path):
        return {"ok": False}

    wb = load_workbook(path)
    ws = wb.active

    fila_objetivo = None

    for fila in range(1, ws.max_row + 1):

        valor = str(
            ws.cell(fila, 1).value or ""
        ).strip()

        if valor == c.archivo:

            fila_objetivo = fila

    if fila_objetivo is None:

        print("NO SE HA ENCONTRADO LA FACTURA")
        return {"ok": False}

    headers = None

    for fila in range(fila_objetivo - 1, 0, -1):

        if str(ws.cell(fila, 1).value or "").strip() == "Archivo":

            headers = fila
            break

    if headers is None:

        print("NO SE HA ENCONTRADO LA CABECERA")
        return {"ok": False}

    columna_objetivo = None

    for col in range(1, ws.max_column + 1):

        nombre = str(
            ws.cell(headers, col).value or ""
        ).strip()

        if nombre == c.columna:

            columna_objetivo = col
            break

    if columna_objetivo is None:

        print("NO SE HA ENCONTRADO LA COLUMNA")
        return {"ok": False}

    ws.cell(
        fila_objetivo,
        columna_objetivo
    ).value = c.valor_nuevo

    print(
        f"ACTUALIZADA FILA {fila_objetivo} "
        f"COLUMNA {columna_objetivo}"
    )

    wb.save(path)

    print("CAMBIOS GUARDADOS")

    guardar_factura_corregida(
    archivo=c.archivo,
    columna=c.columna,
    valor=c.valor_nuevo,
    usuario=c.usuario
)
    wb.save(path)

    print("CAMBIOS GUARDADOS")

    guardar_factura_corregida_completa(
            fila_completa=c.fila_completa,
            usuario=c.usuario
        )

    return {"ok": True}
