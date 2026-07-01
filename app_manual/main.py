import os
import sys
from datetime import datetime
from typing import List

# Garantiza que 'logic' se encuentre aunque uvicorn se lance desde el directorio raíz
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel

from logic import (
    CORREGIR_DIR,
    HISTORIAL_DIR,
    EXPECTED_HEADERS,
    read_pdf_text,
    extract_invoice_with_agent,
    combinar_csvs,
    csv_to_matrix,
    guardar_historial,
    cargar_pdf_pendiente_individual,
    confirmar_y_mover_factura,
)

app = FastAPI(title="Corrección manual de facturas")

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_APP_DIR)

TEMPLATES_DIR = os.path.join(_APP_DIR, "templates")
STATIC_DIR = os.path.join(_PROJECT_ROOT, "static")
UPLOADS_DIR = os.path.join(_PROJECT_ROOT, "uploads")

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================================================
# MODELOS
# =========================================================

class ConfirmacionFactura(BaseModel):
    archivo: str
    fila_completa: list
    usuario: str


# =========================================================
# HOME
# =========================================================

@app.get("/")
def home():
    return FileResponse(os.path.join(TEMPLATES_DIR, "index.html"))


# =========================================================
# PENDIENTES (facturas en corregir_manualmente)
# =========================================================

@app.get("/pendientes-lista")
def pendientes_lista():
    if not os.path.exists(CORREGIR_DIR):
        return {"archivos": []}
    archivos = sorted(
        f for f in os.listdir(CORREGIR_DIR)
        if f.lower().endswith(".pdf")
    )
    return {"archivos": archivos}


@app.post("/extraer-pendiente")
async def extraer_pendiente(body: dict):
    archivo = str(body.get("archivo", "")).strip()
    if not archivo:
        raise HTTPException(status_code=400, detail="Falta el nombre del archivo.")

    fila, fuente = cargar_pdf_pendiente_individual(archivo)

    if fila is None:
        raise HTTPException(status_code=500, detail="No se pudieron extraer datos del PDF.")

    return {"tabla": [EXPECTED_HEADERS, fila], "fuente": fuente}


@app.post("/confirmar-factura")
async def confirmar_factura(body: ConfirmacionFactura):
    try:
        confirmar_y_mover_factura(
            archivo=body.archivo,
            fila_completa=body.fila_completa,
            usuario=body.usuario,
        )
        return {"ok": True}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"No se encontró el PDF: {body.archivo}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =========================================================
# EXTRACCIÓN DE PDF SUBIDO MANUALMENTE (flujo secundario)
# =========================================================

@app.post("/extraer")
async def extraer(facturas: List[UploadFile] = File(...)):
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    results = []

    for f in facturas:
        ruta_temp = os.path.join(UPLOADS_DIR, f.filename)
        content = await f.read()

        with open(ruta_temp, "wb") as fp:
            fp.write(content)

        try:
            text = read_pdf_text(ruta_temp)
            result_csv = extract_invoice_with_agent(
                file_name=f.filename,
                invoice_text=text,
            )
            results.append(result_csv)
            guardar_historial(result_csv, "usuario")
        except Exception as e:
            print(f"ERROR {f.filename}: {e}")
        finally:
            try:
                os.remove(ruta_temp)
            except Exception:
                pass

    if not results:
        raise HTTPException(status_code=500, detail="No se pudieron extraer datos.")

    combined = combinar_csvs(results)
    tabla = csv_to_matrix(combined)
    return {"tabla": tabla}


# =========================================================
# HISTORIAL
# =========================================================

@app.get("/historial-json")
def obtener_historial():
    fecha = datetime.now().strftime("%Y-%m-%d")
    path = os.path.join(HISTORIAL_DIR, f"historial_facturas_usuario_{fecha}.xlsx")

    if not os.path.exists(path):
        return {"tabla": []}

    wb = load_workbook(path)
    ws = wb.active
    tabla = [
        [str(c) if c is not None else "" for c in row]
        for row in ws.iter_rows(values_only=True)
    ]
    return {"tabla": tabla}


@app.get("/historial")
def descargar_historial():
    fecha = datetime.now().strftime("%Y-%m-%d")
    path = os.path.join(HISTORIAL_DIR, f"historial_facturas_usuario_{fecha}.xlsx")

    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="No existe historial para hoy.")

    return FileResponse(
        path,
        filename=f"historial_facturas_{fecha}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
